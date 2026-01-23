import openpyxl
from decimal import Decimal

wb = openpyxl.load_workbook(r'C:\Users\patricio dunstan sae\GastosVsProduccion\informe_produccion_gastos.xlsx')

def parse_valor(valor):
    if valor is None:
        return Decimal('0')
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    if isinstance(valor, str):
        limpio = valor.replace('.', '').replace(',', '')
        if limpio.isdigit():
            return Decimal(limpio)
    return Decimal('0')

print('=== ANÁLISIS COMPLETO DE LOS DATOS DEL EXCEL ===')

# 1. REPUESTOS (DATABODEGA)
print('\n1. REPUESTOS DE DATABODEGA (Q4 2025):')
ws = wb['Detalle Gastos Completo']

total_repuestos = Decimal('0')
maquinas_con_repuestos = set()

for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0]:
        repuestos = parse_valor(row[2])
        if repuestos > 0:
            maquinas_con_repuestos.add(row[0])
            total_repuestos += repuestos

print(f'   Total repuestos: ${total_repuestos:,.0f}')
print(f'   Máquinas con repuestos: {len(maquinas_con_repuestos)}')

# 2. GASTOS OPERACIONALES
print('\n2. GASTOS OPERACIONALES (REPORTES CONTABLES):')
print('   Solo considerando cuentas de gastos (401xxx)')

categorias = {
    'combustibles': 3,
    'reparaciones': 4,
    'seguros': 5,
    'honorarios': 6,
    'epp': 7,
    'peajes': 8,
    'remuneraciones': 9,
    'permisos': 10,
    'alimentacion': 11,
    'pasajes': 12,
    'correspondencia': 13,
    'gastos_legales': 14,
    'multas': 15,
    'otros_gastos': 16
}

total_gastos_op = Decimal('0')
for cat, col_idx in categorias.items():
    categoria_total = Decimal('0')
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0]:
            monto = parse_valor(row[col_idx])
            categoria_total += monto
    print(f'   {cat:20}: ${categoria_total:,.0f}')
    total_gastos_op += categoria_total

print(f'\n   TOTAL GASTOS OPERACIONALES: ${total_gastos_op:,.0f}')

# 3. HORAS HOMBRE
print('\n3. VERIFICACIÓN DE HORAS HOMBRE Y COSTO HH:')
from src.infrastructure.csv.HorasHombreCSVReader import HorasHombreCSVReader

reader_hh = HorasHombreCSVReader(r'C:\Users\patricio dunstan sae\GastosVsProduccion\gastos\_Harcha Maquinaria- HH_Copia de MAQVSOTSVSHH_Tabla.csv')
hh = reader_hh.leer()

total_horas = sum(h.horas for h in hh)
total_costo_hh = total_horas * Decimal('35000')

print(f'   Total horas: {total_horas:,.2f}')
print(f'   Costo HH (@$35.000): ${total_costo_hh:,.0f}')

# 4. RESUMEN GENERAL
print('\n=== RESUMEN GENERAL DE GASTOS (Q4 2025) ===')
print(f'   Repuestos (DATABODEGA):       ${total_repuestos:,.0f}')
print(f'   Costo Horas Hombre:           ${total_costo_hh:,.0f}')
print(f'   Gastos Operacionales (401xxx): ${total_gastos_op:,.0f}')
print('-' * 60)
total_general = total_repuestos + total_costo_hh + total_gastos_op
print(f'   TOTAL GENERAL GASTOS:          ${total_general:,.0f}')

print('\n=== COMPARACIÓN CON VALORES ESPERADOS DEL ANÁLISIS ANTERIOR ===')
print('   Valores reportados anteriormente:')
print('   - Repuestos Q4 2025:     $171.189.015')
print('   - Gastos operacionales:    $700.962.517 (filtrando solo 401xxx)')
print(f'\n   Valores en el Excel:')
print(f'   - Repuestos:              ${total_repuestos:,.0f}')
print(f'   - Gastos operacionales:    ${total_gastos_op:,.0f}')
print(f'\n   Diferencia Repuestos:      ${total_repuestos - Decimal("171189015"):,.0f}')
print(f'   Diferencia Gastos Op:     ${total_gastos_op - Decimal("700962517"):,.0f}')
