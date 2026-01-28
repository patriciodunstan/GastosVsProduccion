"""
Exportador de datos a Excel.

Genera un archivo Excel con múltiples hojas conteniendo los informes.
Sigue el principio de responsabilidad única (SRP).
"""

from decimal import Decimal
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.entities.Produccion import Produccion
from src.domain.entities.HorasHombre import HorasHombre
from src.domain.entities.Repuesto import Repuesto
from src.domain.entities.Leasing import Leasing
from src.domain.entities.GastoOperacional import GastoOperacional
from src.domain.services.CalculadorProduccionReal import CalculadorProduccionReal
from src.domain.services.CalculadorGastos import CalculadorGastos


class ExcelExporter:
    """
    Exporta los datos del informe a un archivo Excel.
    
    Crea un archivo con múltiples hojas:
    - Resumen Trimestral
    - Detalle Producción Mensual
    - Detalle Gastos Mensual
    - Desglose Repuestos
    - Desglose Horas Hombre
    """
    
    MESES = {
        10: 'Octubre',
        11: 'Noviembre',
        12: 'Diciembre'
    }
    
    def __init__(self, ruta_salida: str):
        """
        Inicializa el exportador.
        
        Args:
            ruta_salida: Ruta donde se guardará el archivo Excel
        """
        self.ruta_salida = Path(ruta_salida)
        self.workbook = openpyxl.Workbook()
        self._configurar_estilos()
    
    def _configurar_estilos(self):
        """Configura los estilos que se usarán en el Excel."""
        self.estilo_titulo = Font(bold=True, size=14)
        self.estilo_encabezado = Font(bold=True, size=11)
        self.estilo_moneda = Font(size=10)
        self.estilo_numero = Font(size=10)
        
        self.relleno_encabezado = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        self.relleno_encabezado_font = Font(bold=True, color='FFFFFF', size=11)
        
        self.borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _formatear_moneda(self, valor: Decimal) -> str:
        """Formatea un valor como moneda chilena."""
        return f"${valor:,.0f}".replace(',', '.')
    
    def _formatear_numero(self, valor: Decimal, decimales: int = 2) -> str:
        """Formatea un número con decimales."""
        return f"{valor:,.{decimales}f}".replace(',', '.')
    
    def _aplicar_estilo_encabezado(self, celda):
        """Aplica estilo de encabezado a una celda."""
        celda.font = self.relleno_encabezado_font
        celda.fill = self.relleno_encabezado
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border = self.borde
    
    def _aplicar_borde(self, celda):
        """Aplica borde a una celda."""
        celda.border = self.borde
    
    def exportar(
        self,
        producciones: List[Produccion],
        repuestos: List[Repuesto],
        horas_hombre: List[HorasHombre],
        leasing: Optional[List[Leasing]] = None
    ):
        """
        Exporta todos los datos a Excel (solo producción, repuestos, HH, leasing).
        
        Args:
            producciones: Lista de producciones
            repuestos: Lista de repuestos (DATABODEGA)
            horas_hombre: Lista de horas hombre
            leasing: Lista de leasing (opcional)
        """
        # Calcular datos agregados
        datos = CalculadorProduccionReal.calcular_por_maquina_mes(
            producciones, repuestos, horas_hombre, leasing or []
        )
        
        # Crear hojas
        self._crear_hoja_resumen(datos)
        self._crear_hoja_detalle_produccion(datos)
        self._crear_hoja_detalle_gastos(datos)
        self._crear_hoja_desglose_repuestos(repuestos)
        self._crear_hoja_desglose_horas_hombre(horas_hombre)
        
        # Eliminar hoja por defecto si existe
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Guardar archivo
        self.workbook.save(self.ruta_salida)
    
    def exportar_completo(
        self,
        producciones: List[Produccion],
        repuestos: List[Repuesto],
        horas_hombre: List[HorasHombre],
        gastos_operacionales: List[GastoOperacional],
        leasing: Optional[List[Leasing]] = None
    ):
        """
        Exporta todos los datos completos a Excel (producción + gastos operacionales).
        
        Args:
            producciones: Lista de producciones
            repuestos: Lista de repuestos (DATABODEGA)
            horas_hombre: Lista de horas hombre
            gastos_operacionales: Lista de gastos de reportes contables
            leasing: Lista de leasing (opcional)
        """
        # Calcular datos de producción
        datos_produccion = CalculadorProduccionReal.calcular_por_maquina_mes(
            producciones, repuestos, horas_hombre, leasing or []
        )
        
        # Calcular datos de gastos completos
        datos_gastos = CalculadorGastos.calcular_por_maquina_mes_completo(
            repuestos, horas_hombre, gastos_operacionales, leasing or []
        )
        
        # Combinar ambos conjuntos de datos
        datos = self._combinar_datos_produccion_gastos(datos_produccion, datos_gastos)
        
        # Crear hojas
        self._crear_hoja_resumen_completo(datos, producciones)
        self._crear_hoja_detalle_produccion_completo(datos, producciones)
        self._crear_hoja_detalle_gastos_completo(datos)
        self._crear_hoja_desglose_repuestos(repuestos)
        self._crear_hoja_desglose_horas_hombre(horas_hombre)
        self._crear_hoja_desglose_gastos_operacionales(gastos_operacionales)
        
        # Eliminar hoja por defecto si existe
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Guardar archivo
        self.workbook.save(self.ruta_salida)
    
    def _combinar_datos_produccion_gastos(
        self, 
        datos_produccion: Dict[Tuple[str, int], Dict], 
        datos_gastos: Dict[Tuple[str, int], Dict]
    ) -> Dict[Tuple[str, int], Dict]:
        """Combina datos de producción y gastos en una sola estructura."""
        datos_combinados: Dict[Tuple[str, int], Dict] = {}
        
        # Obtener todas las claves únicas
        todas_claves = set(datos_produccion.keys()) | set(datos_gastos.keys())
        
        for clave in todas_claves:
            datos_combinados[clave] = {}
            
            # Agregar datos de producción si existen
            if clave in datos_produccion:
                datos_combinados[clave]['produccion'] = datos_produccion[clave]['produccion']
                datos_combinados[clave]['produccion_neta'] = datos_produccion[clave]['produccion_neta']
                datos_combinados[clave]['produccion_real'] = datos_produccion[clave]['produccion_real']
            else:
                # Valores por defecto si no hay datos de producción
                datos_combinados[clave]['produccion'] = {
                    'mt3': Decimal('0'),
                    'horas_trabajadas': Decimal('0'),
                    'kilometros': Decimal('0'),
                    'vueltas': Decimal('0'),
                    'valor_mt3': Decimal('0'),
                    'valor_horas': Decimal('0'),
                    'valor_km': Decimal('0'),
                    'valor_dias': Decimal('0'),
                    'valor_vueltas': Decimal('0')
                }
                datos_combinados[clave]['produccion_neta'] = {
                    'mt3': Decimal('0'),
                    'horas_trabajadas': Decimal('0'),
                    'kilometros': Decimal('0'),
                    'vueltas': Decimal('0'),
                    'valor_monetario': Decimal('0')
                }
                datos_combinados[clave]['produccion_real'] = {
                    'mt3': Decimal('0'),
                    'horas_trabajadas': Decimal('0'),
                    'kilometros': Decimal('0'),
                    'vueltas': Decimal('0'),
                    'valor_monetario': Decimal('0')
                }
            
            # Agregar datos de gastos si existen
            if clave in datos_gastos:
                datos_combinados[clave]['gastos'] = datos_gastos[clave]
            else:
                # Valores por defecto si no hay datos de gastos
                datos_combinados[clave]['gastos'] = datos_gastos.get(clave, {
                    'repuestos': Decimal('0'),
                    'horas_hombre': Decimal('0'),
                    'costo_hh': Decimal('0'),
                    'leasing': Decimal('0'),
                    'combustibles': Decimal('0'),
                    'reparaciones': Decimal('0'),
                    'seguros': Decimal('0'),
                    'honorarios': Decimal('0'),
                    'epp': Decimal('0'),
                    'peajes': Decimal('0'),
                    'remuneraciones': Decimal('0'),
                    'permisos': Decimal('0'),
                    'alimentacion': Decimal('0'),
                    'pasajes': Decimal('0'),
                    'correspondencia': Decimal('0'),
                    'gastos_legales': Decimal('0'),
                    'multas': Decimal('0'),
                    'otros_gastos': Decimal('0'),
                    'total_gastos_operacionales': Decimal('0'),
                    'total': Decimal('0')
                })
        
        return datos_combinados
    
    def _crear_hoja_detalle_gastos_completo(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea 4 hojas de gastos completos: Octubre, Noviembre, Diciembre y Resumen Trimestral."""
        # Crear hoja individual por cada mes
        self._crear_hoja_gastos_completo_mes(datos, 10, "Gastos Octubre Completo")
        self._crear_hoja_gastos_completo_mes(datos, 11, "Gastos Noviembre Completo")
        self._crear_hoja_gastos_completo_mes(datos, 12, "Gastos Diciembre Completo")
        # Crear resumen trimestral
        self._crear_hoja_resumen_gastos_trimestral_completo(datos)

    def _crear_hoja_gastos_completo_mes(self, datos: Dict[Tuple[str, int], Dict], mes: int, nombre_hoja: str):
        """Crea una hoja de gastos completos para un mes específico."""
        hoja = self.workbook.create_sheet(nombre_hoja)

        # Título
        hoja['A1'] = f'GASTOS COMPLETOS {self.MESES[mes].upper()}'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:O1')

        # Nota
        hoja['A2'] = 'NOTA: Incluye TODOS los gastos operacionales de reportes contables + repuestos DATABODEGA. Valores NETOS (sin IVA).'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:O2')

        # Encabezados
        encabezados = [
            'Máquina', 'Repuestos', 'Combustibles', 'Reparaciones', 'Seguros', 'Honorarios',
            'EPP', 'Peajes', 'Remuneraciones', 'Permisos', 'Alimentación', 'Pasajes',
            'Correspondencia', 'Gastos Legales', 'Multas', 'Otros', 'Total Gastos'
        ]

        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)

        # Filtrar datos por mes y ordenar por total de gastos (mayor a menor)
        fila = 5
        datos_mes = sorted(
            [(maq, datos_mes) for (maq, m), datos_mes in datos.items() if m == mes],
            key=lambda x: (
                x[1].get('gastos', {}).get('repuestos', Decimal('0')) +
                x[1].get('gastos', {}).get('horas_hombre', Decimal('0')) * Decimal('35000') +
                x[1].get('gastos', {}).get('leasing', Decimal('0')) +
                x[1].get('gastos', {}).get('total_gastos_operacionales', Decimal('0'))
            ),
            reverse=True
        )

        for maquina, datos_mes in datos_mes:
            gastos = datos_mes.get('gastos', {})

            hoja.cell(row=fila, column=1, value=maquina)
            hoja.cell(row=fila, column=2, value=self._formatear_moneda(gastos.get('repuestos', Decimal('0'))))
            hoja.cell(row=fila, column=3, value=self._formatear_moneda(gastos.get('combustibles', Decimal('0'))))
            hoja.cell(row=fila, column=4, value=self._formatear_moneda(gastos.get('reparaciones', Decimal('0'))))
            hoja.cell(row=fila, column=5, value=self._formatear_moneda(gastos.get('seguros', Decimal('0'))))
            hoja.cell(row=fila, column=6, value=self._formatear_moneda(gastos.get('honorarios', Decimal('0'))))
            hoja.cell(row=fila, column=7, value=self._formatear_moneda(gastos.get('epp', Decimal('0'))))
            hoja.cell(row=fila, column=8, value=self._formatear_moneda(gastos.get('peajes', Decimal('0'))))
            hoja.cell(row=fila, column=9, value=self._formatear_moneda(gastos.get('remuneraciones', Decimal('0'))))
            hoja.cell(row=fila, column=10, value=self._formatear_moneda(gastos.get('permisos', Decimal('0'))))
            hoja.cell(row=fila, column=11, value=self._formatear_moneda(gastos.get('alimentacion', Decimal('0'))))
            hoja.cell(row=fila, column=12, value=self._formatear_moneda(gastos.get('pasajes', Decimal('0'))))
            hoja.cell(row=fila, column=13, value=self._formatear_moneda(gastos.get('correspondencia', Decimal('0'))))
            hoja.cell(row=fila, column=14, value=self._formatear_moneda(gastos.get('gastos_legales', Decimal('0'))))
            hoja.cell(row=fila, column=15, value=self._formatear_moneda(gastos.get('multas', Decimal('0'))))
            hoja.cell(row=fila, column=16, value=self._formatear_moneda(gastos.get('otros_gastos', Decimal('0'))))

            # Calcular total real (repuestos + gastos operacionales + HH + leasing)
            total_real = (
                gastos.get('repuestos', Decimal('0')) +
                gastos.get('horas_hombre', Decimal('0')) * Decimal('35000') +
                gastos.get('leasing', Decimal('0')) +
                gastos.get('total_gastos_operacionales', Decimal('0'))
            )
            hoja.cell(row=fila, column=17, value=self._formatear_moneda(total_real))

            # Estilo para celdas de datos
            for col in range(1, 18):
                self._aplicar_borde(hoja.cell(row=fila, column=col))

            fila += 1

        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        for col in range(2, 18):
            hoja.column_dimensions[get_column_letter(col)].width = 15

    def _crear_hoja_resumen_gastos_trimestral_completo(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea la hoja de resumen trimestral de gastos completos."""
        hoja = self.workbook.create_sheet("Resumen Gastos Trimestral Completo")

        # Título
        hoja['A1'] = 'RESUMEN GASTOS TRIMESTRAL COMPLETO'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:BM1')

        # Nota
        hoja['A2'] = 'NOTA: Incluye TODOS los gastos operacionales de reportes contables + repuestos DATABODEGA. Valores NETOS (sin IVA).'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:BM2')

        # Definir categorías de gastos
        categorias = [
            'Repuestos', 'Combustibles', 'Reparaciones', 'Seguros', 'Honorarios',
            'EPP', 'Peajes', 'Remuneraciones', 'Permisos', 'Alimentación', 'Pasajes',
            'Correspondencia', 'Gastos Legales', 'Multas', 'Otros', 'Total'
        ]

        # Construir encabezados
        encabezados = ['Máquina']
        for mes in [10, 11, 12]:
            nombre_mes = self.MESES[mes]
            for cat in categorias:
                encabezados.append(f'{cat} {nombre_mes[:3]}')  # Oct, Nov, Dic

        # Agregar columnas de totales trimestrales
        for cat in categorias:
            encabezados.append(f'Total {cat}')

        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)

        # Obtener todas las máquinas únicas con su total general y ordenar por total (mayor a menor)
        maquinas_con_total = []
        for maquina in set(maq for maq, _ in datos.keys()):
            total_general = Decimal('0')
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    gastos = datos[clave].get('gastos', {})
                    total_general += (
                        gastos.get('repuestos', Decimal('0')) +
                        gastos.get('horas_hombre', Decimal('0')) * Decimal('35000') +
                        gastos.get('leasing', Decimal('0')) +
                        gastos.get('total_gastos_operacionales', Decimal('0'))
                    )
            maquinas_con_total.append((maquina, total_general))

        # Ordenar por total general (mayor a menor)
        maquinas_ordenadas = sorted(maquinas_con_total, key=lambda x: x[1], reverse=True)

        fila = 5
        for maquina, _ in maquinas_ordenadas:
            col = 1
            hoja.cell(row=fila, column=col, value=maquina)
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1

            # Inicializar acumuladores para totales trimestrales
            totales_trimestral = {cat: Decimal('0') for cat in categorias}

            # Datos por cada mes
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    gastos = datos[clave].get('gastos', {})

                    # Valores del mes
                    valores_mes = {
                        'Repuestos': gastos.get('repuestos', Decimal('0')),
                        'Combustibles': gastos.get('combustibles', Decimal('0')),
                        'Reparaciones': gastos.get('reparaciones', Decimal('0')),
                        'Seguros': gastos.get('seguros', Decimal('0')),
                        'Honorarios': gastos.get('honorarios', Decimal('0')),
                        'EPP': gastos.get('epp', Decimal('0')),
                        'Peajes': gastos.get('peajes', Decimal('0')),
                        'Remuneraciones': gastos.get('remuneraciones', Decimal('0')),
                        'Permisos': gastos.get('permisos', Decimal('0')),
                        'Alimentación': gastos.get('alimentacion', Decimal('0')),
                        'Pasajes': gastos.get('pasajes', Decimal('0')),
                        'Correspondencia': gastos.get('correspondencia', Decimal('0')),
                        'Gastos Legales': gastos.get('gastos_legales', Decimal('0')),
                        'Multas': gastos.get('multas', Decimal('0')),
                        'Otros': gastos.get('otros_gastos', Decimal('0')),
                        'Total': (
                            gastos.get('repuestos', Decimal('0')) +
                            gastos.get('horas_hombre', Decimal('0')) * Decimal('35000') +
                            gastos.get('leasing', Decimal('0')) +
                            gastos.get('total_gastos_operacionales', Decimal('0'))
                        )
                    }

                    # Escribir valores del mes
                    for cat in categorias:
                        valor = valores_mes[cat]
                        hoja.cell(row=fila, column=col, value=self._formatear_moneda(valor))
                        self._aplicar_borde(hoja.cell(row=fila, column=col))
                        totales_trimestral[cat] += valor
                        col += 1
                else:
                    # Sin datos para este mes
                    for _ in categorias:
                        hoja.cell(row=fila, column=col, value='-')
                        self._aplicar_borde(hoja.cell(row=fila, column=col))
                        col += 1

            # Escribir totales trimestrales
            for cat in categorias:
                hoja.cell(row=fila, column=col, value=self._formatear_moneda(totales_trimestral[cat]))
                self._aplicar_borde(hoja.cell(row=fila, column=col))
                col += 1

            fila += 1

        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        for col in range(2, len(encabezados) + 1):
            hoja.column_dimensions[get_column_letter(col)].width = 13
    
    def _crear_hoja_detalle_produccion_completo(self, datos: Dict[Tuple[str, int], Dict], producciones: List[Produccion]):
        """Crea la hoja de detalle de producción mensual con datos completos."""
        hoja = self.workbook.create_sheet("Detalle Producción Completo")
        
        # Título
        hoja['A1'] = 'DETALLE PRODUCCIÓN MENSUAL COMPLETO'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:G1')
        
        # Nota
        hoja['A2'] = 'NOTA: Datos de producción combinados con gastos operacionales. Valores NETOS (sin IVA).'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:G2')
        
        # Encabezados
        encabezados = [
            'Máquina', 'Mes', 'MT3', 'Horas Trabajadas', 'Kilómetros', 'Vueltas', 'Producción Real ($)'
        ]
        
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        # Datos por (maquina, mes)
        fila = 5
        for (maquina, mes), datos_mes in sorted(datos.items()):
            prod = datos_mes['produccion']
            
            hoja.cell(row=fila, column=1, value=maquina)
            hoja.cell(row=fila, column=2, value=self.MESES[mes])
            hoja.cell(row=fila, column=3, value=float(prod['mt3']))
            hoja.cell(row=fila, column=4, value=float(prod['horas_trabajadas']))
            hoja.cell(row=fila, column=5, value=float(prod['kilometros']))
            hoja.cell(row=fila, column=6, value=float(prod['vueltas']))
            hoja.cell(row=fila, column=7, value=self._formatear_moneda(datos_mes['produccion_real']['valor_monetario']))
            
            # Estilo para celdas de datos
            for col in range(1, 8):
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        hoja.column_dimensions['B'].width = 15
        for col in range(3, 8):
            hoja.column_dimensions[get_column_letter(col)].width = 18
    
    def _crear_hoja_resumen_completo(self, datos: Dict[Tuple[str, int], Dict], producciones: List[Produccion]):
        """Crea la hoja de resumen trimestral con desglose completo."""
        hoja = self.workbook.create_sheet("Resumen Trimestral Completo")
        
        # Título
        hoja['A1'] = 'RESUMEN TRIMESTRAL - PRODUCCIÓN VS GASTOS COMPLETO'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:P1')
        
        # Nota
        hoja['A2'] = 'NOTA: Incluye TODOS los gastos operacionales de reportes contables + repuestos DATABODEGA + datos de producción. Valores NETOS (sin IVA).'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:P2')
        
        # Encabezados
        encabezados = [
            'Máquina', 'Producción (MT3)', 'Horas (H)', 'Km', 'Vueltas', 'Val. Prod. Neto',
            'Repuestos', 'Combustibles', 'Reparaciones', 'Seguros', 'Honorarios',
            'EPP', 'Peajes', 'Remuneraciones', 'Permisos', 'Alimentación', 'Pasajes',
            'Correspondencia', 'Gastos Legales', 'Multas', 'Otros', 'Total Op.', 'Total Prod. Neto'
        ]
        
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        # Obtener todas las máquinas únicas
        maquinas = set()
        for (maquina, mes) in datos.keys():
            maquinas.add(maquina)
        
        fila = 5
        for maquina in sorted(maquinas):
            # Calcular totales por máquina (trimestral)
            total_prod_mt3 = Decimal('0')
            total_prod_h = Decimal('0')
            total_prod_km = Decimal('0')
            total_prod_vueltas = Decimal('0')
            
            total_repuestos = Decimal('0')
            total_combustibles = Decimal('0')
            total_reparaciones = Decimal('0')
            total_seguros = Decimal('0')
            total_honorarios = Decimal('0')
            total_epp = Decimal('0')
            total_peajes = Decimal('0')
            total_remuneraciones = Decimal('0')
            total_permisos = Decimal('0')
            total_alimentacion = Decimal('0')
            total_pasajes = Decimal('0')
            total_correspondencia = Decimal('0')
            total_gastos_legales = Decimal('0')
            total_multas = Decimal('0')
            total_otros = Decimal('0')
            total_gastos_op = Decimal('0')
            
            total_prod_neta = Decimal('0')
            total_prod_real = Decimal('0')
            
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    d = datos[clave]['produccion']
                    g = datos[clave]['gastos']
                    
                    total_prod_mt3 += d['mt3']
                    total_prod_h += d['horas_trabajadas']
                    total_prod_km += d['kilometros']
                    total_prod_vueltas += d['vueltas']
                    
                    total_repuestos += g.get('repuestos', Decimal('0'))
                    total_combustibles += g.get('combustibles', Decimal('0'))
                    total_reparaciones += g.get('reparaciones', Decimal('0'))
                    total_seguros += g.get('seguros', Decimal('0'))
                    total_honorarios += g.get('honorarios', Decimal('0'))
                    total_epp += g.get('epp', Decimal('0'))
                    total_peajes += g.get('peajes', Decimal('0'))
                    total_remuneraciones += g.get('remuneraciones', Decimal('0'))
                    total_permisos += g.get('permisos', Decimal('0'))
                    total_alimentacion += g.get('alimentacion', Decimal('0'))
                    total_pasajes += g.get('pasajes', Decimal('0'))
                    total_correspondencia += g.get('correspondencia', Decimal('0'))
                    total_gastos_legales += g.get('gastos_legales', Decimal('0'))
                    total_multas += g.get('multas', Decimal('0'))
                    total_otros += g.get('otros_gastos', Decimal('0'))
                    total_gastos_op += g.get('total_gastos_operacionales', Decimal('0'))
                    
                    total_prod_neta += datos[clave]['produccion_neta']['valor_monetario']
                    total_prod_real += datos[clave]['produccion_real']['valor_monetario']
            
            # Totales
            total_gastos = (
                total_repuestos +
                total_combustibles +
                total_reparaciones +
                total_seguros +
                total_honorarios +
                total_epp +
                total_peajes +
                total_remuneraciones +
                total_permisos +
                total_alimentacion +
                total_pasajes +
                total_correspondencia +
                total_gastos_legales +
                total_multas +
                total_otros +
                total_gastos_op
            )
            
            # Escribir fila
            hoja.cell(row=fila, column=1, value=maquina)
            hoja.cell(row=fila, column=2, value=self._formatear_numero(total_prod_mt3, 0))
            hoja.cell(row=fila, column=3, value=self._formatear_numero(total_prod_h, 2))
            hoja.cell(row=fila, column=4, value=self._formatear_numero(total_prod_km, 0))
            hoja.cell(row=fila, column=5, value=self._formatear_numero(total_prod_vueltas, 0))
            
            hoja.cell(row=fila, column=6, value=self._formatear_moneda(total_prod_neta))
            hoja.cell(row=fila, column=7, value=self._formatear_moneda(total_prod_real))
            
            hoja.cell(row=fila, column=8, value=self._formatear_moneda(total_repuestos))
            hoja.cell(row=fila, column=9, value=self._formatear_moneda(total_combustibles))
            hoja.cell(row=fila, column=10, value=self._formatear_moneda(total_reparaciones))
            hoja.cell(row=fila, column=11, value=self._formatear_moneda(total_seguros))
            hoja.cell(row=fila, column=12, value=self._formatear_moneda(total_honorarios))
            hoja.cell(row=fila, column=13, value=self._formatear_moneda(total_epp))
            hoja.cell(row=fila, column=14, value=self._formatear_moneda(total_peajes))
            hoja.cell(row=fila, column=15, value=self._formatear_moneda(total_remuneraciones))
            hoja.cell(row=fila, column=16, value=self._formatear_moneda(total_permisos))
            hoja.cell(row=fila, column=17, value=self._formatear_moneda(total_alimentacion))
            hoja.cell(row=fila, column=18, value=self._formatear_moneda(total_pasajes))
            hoja.cell(row=fila, column=19, value=self._formatear_moneda(total_correspondencia))
            hoja.cell(row=fila, column=20, value=self._formatear_moneda(total_gastos_legales))
            hoja.cell(row=fila, column=21, value=self._formatear_moneda(total_multas))
            hoja.cell(row=fila, column=22, value=self._formatear_moneda(total_otros))
            hoja.cell(row=fila, column=23, value=self._formatear_moneda(total_gastos_op))
            
            # Total parcial (sin repuestos ni horas hombre ni leasing)
            total_parcial = (
                total_combustibles +
                total_reparaciones +
                total_seguros +
                total_honorarios +
                total_epp +
                total_peajes +
                total_remuneraciones +
                total_permisos +
                total_alimentacion +
                total_pasajes +
                total_correspondencia +
                total_gastos_legales +
                total_multas +
                total_otros
            )
            
            hoja.cell(row=fila, column=24, value=self._formatear_moneda(total_parcial))
            
            # Total general
            hoja.cell(row=fila, column=25, value=self._formatear_moneda(total_gastos))
            
            # Total neto (producción real - total gastos)
            resultado = total_prod_real - total_gastos
            hoja.cell(row=fila, column=26, value=self._formatear_moneda(resultado))
            
            # Estilo de resultado
            celda_resultado = hoja.cell(row=fila, column=26)
            celda_resultado.alignment = Alignment(horizontal='center', vertical='center')
            if resultado > 0:
                celda_resultado.font = Font(color='28a745', bold=True)
            else:
                celda_resultado.font = Font(color='dc3545', bold=True)
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 30
        for col in range(2, 27):
            hoja.column_dimensions[get_column_letter(col)].width = 14
        
        # Ajustar columna de resultado
        hoja.column_dimensions['AA'].width = 18
        for col in range(28, 31):
            hoja.column_dimensions[get_column_letter(col)].width = 18
    
    def _crear_hoja_desglose_gastos_operacionales(self, gastos_operacionales: List[GastoOperacional]):
        """Crea una hoja con desglose detallado de gastos operacionales."""
        hoja = self.workbook.create_sheet("Desglose Gastos por Tipo")
        
        # Título
        hoja['A1'] = 'DESGLOSE DE GASTOS POR TIPO Y MÁQUINA'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:G1')
        
        # Encabezados
        encabezados = [
            'Código Máquina', 'Mes', 'Tipo Gasto', 'Nombre Tipo Gasto', 'Glosa', 'Origen', 'Monto'
        ]
        
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=3, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        # Ordenar por máquina y fecha
        fila = 4
        for gasto in sorted(gastos_operacionales, key=lambda g: (g.codigo_maquina, g.fecha)):
            hoja.cell(row=fila, column=1, value=gasto.codigo_maquina)
            hoja.cell(row=fila, column=2, value=self.MESES[gasto.fecha.month])
            hoja.cell(row=fila, column=3, value=gasto.tipo_gasto)
            hoja.cell(row=fila, column=4, value=gasto.nombre_tipo_gasto)
            hoja.cell(row=fila, column=5, value=gasto.glosa)
            hoja.cell(row=fila, column=6, value=gasto.origen)
            hoja.cell(row=fila, column=7, value=self._formatear_moneda(gasto.monto))
            
            for col in range(1, 8):
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 20
        for col in range(2, 8):
            hoja.column_dimensions[get_column_letter(col)].width = 18
    
    def _crear_hoja_resumen(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea la hoja de resumen trimestral."""
        hoja = self.workbook.create_sheet("Resumen Trimestral")
        
        # Título
        hoja['A1'] = 'INFORME TRIMESTRAL - PRODUCCIÓN VS GASTOS'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:P1')
        
        # Nota sobre valores netos
        hoja['A2'] = 'NOTA: Todos los valores monetarios son NETOS (sin IVA). Leasing con IVA descontado (19%). Repuestos sin IVA.'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:P2')
        
        # Encabezados
        encabezados = [
            'Máquina', 'Producción Oct', 'Prod. Neta Oct', 'Gastos Oct', 'Prod. Real Oct',
            'Producción Nov', 'Prod. Neta Nov', 'Gastos Nov', 'Prod. Real Nov',
            'Producción Dic', 'Prod. Neta Dic', 'Gastos Dic', 'Prod. Real Dic',
            'Total Producción', 'Total Prod. Neta', 'Total Gastos', 'Total Prod. Real'
        ]
        
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        # Obtener todas las máquinas únicas y calcular producción real total por máquina
        maquinas_con_total = []
        for maquina in set(codigo for codigo, _ in datos.keys()):
            total_prod_real = Decimal('0')
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    total_prod_real += datos[clave]['produccion_real']['valor_monetario']
            maquinas_con_total.append((maquina, total_prod_real))
        
        # Ordenar por producción real (de menor a mayor)
        maquinas_ordenadas = sorted(maquinas_con_total, key=lambda x: x[1])
        
        fila = 5
        for maquina, _ in maquinas_ordenadas:
            # Calcular totales por máquina
            total_prod = {'mt3': Decimal('0'), 'horas_trabajadas': Decimal('0'),
                         'kilometros': Decimal('0'), 'vueltas': Decimal('0')}
            total_prod_neta = Decimal('0')
            total_gastos = Decimal('0')
            total_prod_real = Decimal('0')
            
            # Datos por mes
            datos_mes = {}
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    datos_mes[mes] = datos[clave]
                    prod = datos[clave]['produccion']
                    prod_neta = datos[clave]['produccion_neta']
                    gastos = datos[clave]['gastos']
                    prod_real = datos[clave]['produccion_real']
                    
                    total_prod['mt3'] += prod['mt3']
                    total_prod['horas_trabajadas'] += prod['horas_trabajadas']
                    total_prod['kilometros'] += prod['kilometros']
                    total_prod['vueltas'] += prod['vueltas']
                    total_prod_neta += prod_neta['valor_monetario']
                    total_gastos += gastos['total']
                    total_prod_real += prod_real['valor_monetario']
            
            # Escribir datos
            col = 1
            hoja.cell(row=fila, column=col, value=maquina)
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            for mes in [10, 11, 12]:
                col += 1
                if mes in datos_mes:
                    prod = datos_mes[mes]['produccion']
                    prod_str = f"MT3:{prod['mt3']:.0f} H:{prod['horas_trabajadas']:.0f} KM:{prod['kilometros']:.0f}"
                    hoja.cell(row=fila, column=col, value=prod_str)
                else:
                    hoja.cell(row=fila, column=col, value='-')
                self._aplicar_borde(hoja.cell(row=fila, column=col))
                
                col += 1
                if mes in datos_mes:
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(datos_mes[mes]['produccion_neta']['valor_monetario']))
                else:
                    hoja.cell(row=fila, column=col, value='-')
                self._aplicar_borde(hoja.cell(row=fila, column=col))
                
                col += 1
                if mes in datos_mes:
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(datos_mes[mes]['gastos']['total']))
                else:
                    hoja.cell(row=fila, column=col, value='-')
                self._aplicar_borde(hoja.cell(row=fila, column=col))
                
                col += 1
                if mes in datos_mes:
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(datos_mes[mes]['produccion_real']['valor_monetario']))
                else:
                    hoja.cell(row=fila, column=col, value='-')
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            # Totales
            col += 1
            total_prod_str = f"MT3:{total_prod['mt3']:.0f} H:{total_prod['horas_trabajadas']:.0f} KM:{total_prod['kilometros']:.0f}"
            hoja.cell(row=fila, column=col, value=total_prod_str)
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_prod_neta))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_gastos))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_prod_real))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        for col in range(2, 17):
            hoja.column_dimensions[get_column_letter(col)].width = 18
    
    def _crear_hoja_detalle_produccion(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea la hoja de detalle de producción mensual."""
        hoja = self.workbook.create_sheet("Detalle Producción")
        
        # Título
        hoja['A1'] = 'DETALLE PRODUCCIÓN MENSUAL'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:F1')
        
        # Nota sobre valores netos
        hoja['A2'] = 'NOTA: Valores de producción son NETOS (sin IVA)'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:F2')
        
        # Encabezados
        encabezados = ['Máquina', 'Mes', 'MT3', 'Horas Trabajadas', 'Kilómetros', 'Vueltas']
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        fila = 5
        for (maquina, mes), datos_mes in sorted(datos.items()):
            prod = datos_mes['produccion']
            
            hoja.cell(row=fila, column=1, value=maquina)
            hoja.cell(row=fila, column=2, value=self.MESES[mes])
            hoja.cell(row=fila, column=3, value=float(prod['mt3']))
            hoja.cell(row=fila, column=4, value=float(prod['horas_trabajadas']))
            hoja.cell(row=fila, column=5, value=float(prod['kilometros']))
            hoja.cell(row=fila, column=6, value=float(prod['vueltas']))
            
            for col in range(1, 7):
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        hoja.column_dimensions['B'].width = 15
        for col in range(3, 7):
            hoja.column_dimensions[get_column_letter(col)].width = 18
    
    def _crear_hoja_detalle_gastos(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea 4 hojas de gastos: Octubre, Noviembre, Diciembre y Resumen Trimestral."""
        # Crear hoja individual por cada mes
        self._crear_hoja_gastos_mes(datos, 10, "Gastos Octubre")
        self._crear_hoja_gastos_mes(datos, 11, "Gastos Noviembre")
        self._crear_hoja_gastos_mes(datos, 12, "Gastos Diciembre")
        # Crear resumen trimestral
        self._crear_hoja_resumen_gastos_trimestral(datos)

    def _crear_hoja_gastos_mes(self, datos: Dict[Tuple[str, int], Dict], mes: int, nombre_hoja: str):
        """Crea una hoja de gastos para un mes específico."""
        hoja = self.workbook.create_sheet(nombre_hoja)

        # Título
        hoja['A1'] = f'GASTOS {self.MESES[mes].upper()}'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:F1')

        # Nota sobre valores netos
        hoja['A2'] = 'NOTA: Valores NETOS (sin IVA). Leasing: IVA descontado. Repuestos: sin IVA.'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:F2')

        # Encabezados
        encabezados = ['Máquina', 'Repuestos', 'Horas Hombre', 'Costo HH', 'Leasing', 'Total Gastos']
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)

        # Filtrar datos por mes y ordenar por total de gastos (mayor a menor)
        fila = 5
        datos_mes = sorted(
            [(maq, datos_mes) for (maq, m), datos_mes in datos.items() if m == mes],
            key=lambda x: x[1]['gastos']['total'],
            reverse=True
        )

        for maquina, datos_mes in datos_mes:
            gastos = datos_mes['gastos']

            hoja.cell(row=fila, column=1, value=maquina)
            hoja.cell(row=fila, column=2, value=self._formatear_moneda(gastos['repuestos']))
            hoja.cell(row=fila, column=3, value=float(gastos['horas_hombre']))
            hoja.cell(row=fila, column=4, value=self._formatear_moneda(gastos['costo_hh']))
            hoja.cell(row=fila, column=5, value=self._formatear_moneda(gastos.get('leasing', Decimal('0'))))
            hoja.cell(row=fila, column=6, value=self._formatear_moneda(gastos['total']))

            for col in range(1, 7):
                self._aplicar_borde(hoja.cell(row=fila, column=col))

            fila += 1

        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        for col in range(2, 7):
            hoja.column_dimensions[get_column_letter(col)].width = 18

    def _crear_hoja_resumen_gastos_trimestral(self, datos: Dict[Tuple[str, int], Dict]):
        """Crea la hoja de resumen trimestral de gastos."""
        hoja = self.workbook.create_sheet("Resumen Gastos Trimestral")

        # Título
        hoja['A1'] = 'RESUMEN GASTOS TRIMESTRAL'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:P1')

        # Nota sobre valores netos
        hoja['A2'] = 'NOTA: Valores NETOS (sin IVA). Leasing: IVA descontado. Repuestos: sin IVA.'
        hoja['A2'].font = Font(size=9, italic=True, color='666666')
        hoja.merge_cells('A2:P2')

        # Encabezados
        encabezados = [
            'Máquina',
            'Repuestos Oct', 'HH Oct', 'Costo HH Oct', 'Leasing Oct', 'Total Oct',
            'Repuestos Nov', 'HH Nov', 'Costo HH Nov', 'Leasing Nov', 'Total Nov',
            'Repuestos Dic', 'HH Dic', 'Costo HH Dic', 'Leasing Dic', 'Total Dic',
            'Total Repuestos', 'Total HH', 'Total Costo HH', 'Total Leasing', 'Total General'
        ]
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)

        # Obtener todas las máquinas únicas con su total general y ordenar por total (mayor a menor)
        maquinas_con_total = []
        for maquina in set(maq for maq, _ in datos.keys()):
            total_general = Decimal('0')
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    total_general += datos[clave]['gastos']['total']
            maquinas_con_total.append((maquina, total_general))

        # Ordenar por total general (mayor a menor)
        maquinas_ordenadas = sorted(maquinas_con_total, key=lambda x: x[1], reverse=True)

        fila = 5
        for maquina, _ in maquinas_ordenadas:
            col = 1
            hoja.cell(row=fila, column=col, value=maquina)
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1

            # Acumuladores para totales trimestrales
            total_repuestos = Decimal('0')
            total_hh = Decimal('0')
            total_costo_hh = Decimal('0')
            total_leasing = Decimal('0')
            total_general = Decimal('0')

            # Datos por cada mes
            for mes in [10, 11, 12]:
                clave = (maquina, mes)
                if clave in datos:
                    gastos = datos[clave]['gastos']
                    repuestos = gastos['repuestos']
                    hh = gastos['horas_hombre']
                    costo_hh = gastos['costo_hh']
                    leasing = gastos.get('leasing', Decimal('0'))
                    total = gastos['total']

                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(repuestos))
                    self._aplicar_borde(hoja.cell(row=fila, column=col))
                    col += 1
                    hoja.cell(row=fila, column=col, value=float(hh))
                    self._aplicar_borde(hoja.cell(row=fila, column=col))
                    col += 1
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(costo_hh))
                    self._aplicar_borde(hoja.cell(row=fila, column=col))
                    col += 1
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(leasing))
                    self._aplicar_borde(hoja.cell(row=fila, column=col))
                    col += 1
                    hoja.cell(row=fila, column=col, value=self._formatear_moneda(total))
                    self._aplicar_borde(hoja.cell(row=fila, column=col))
                    col += 1

                    # Acumular totales
                    total_repuestos += repuestos
                    total_hh += hh
                    total_costo_hh += costo_hh
                    total_leasing += leasing
                    total_general += total
                else:
                    # Sin datos para este mes
                    for _ in range(5):
                        hoja.cell(row=fila, column=col, value='-')
                        self._aplicar_borde(hoja.cell(row=fila, column=col))
                        col += 1

            # Escribir totales trimestrales
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_repuestos))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1
            hoja.cell(row=fila, column=col, value=float(total_hh))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_costo_hh))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_leasing))
            self._aplicar_borde(hoja.cell(row=fila, column=col))
            col += 1
            hoja.cell(row=fila, column=col, value=self._formatear_moneda(total_general))
            self._aplicar_borde(hoja.cell(row=fila, column=col))

            fila += 1

        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 25
        for col in range(2, 22):
            hoja.column_dimensions[get_column_letter(col)].width = 15
    
    def _crear_hoja_desglose_repuestos(self, repuestos: List[Repuesto]):
        """Crea la hoja de desglose de repuestos."""
        hoja = self.workbook.create_sheet("Desglose Repuestos")
        
        # Título
        hoja['A1'] = 'DESGLOSE REPUESTOS'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:G1')
        
        # Encabezados
        encabezados = ['Máquina', 'Fecha', 'Repuesto', 'Cantidad', 'Precio Unit.', 'Total', 'Asignado A']
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=3, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        fila = 4
        for repuesto in sorted(repuestos, key=lambda r: (r.codigo_maquina, r.fecha_salida)):
            hoja.cell(row=fila, column=1, value=repuesto.codigo_maquina)
            hoja.cell(row=fila, column=2, value=repuesto.fecha_salida.strftime('%d/%m/%Y'))
            hoja.cell(row=fila, column=3, value=repuesto.nombre)
            hoja.cell(row=fila, column=4, value=float(repuesto.cantidad))
            hoja.cell(row=fila, column=5, value=self._formatear_moneda(repuesto.precio_unitario))
            hoja.cell(row=fila, column=6, value=self._formatear_moneda(repuesto.total))
            hoja.cell(row=fila, column=7, value=repuesto.asignado_a)
            
            for col in range(1, 8):
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 20
        hoja.column_dimensions['B'].width = 12
        hoja.column_dimensions['C'].width = 40
        for col in range(4, 8):
            hoja.column_dimensions[get_column_letter(col)].width = 15
    
    def _crear_hoja_desglose_horas_hombre(self, horas_hombre: List[HorasHombre]):
        """Crea la hoja de desglose de horas hombre."""
        hoja = self.workbook.create_sheet("Desglose Horas Hombre")
        
        # Título
        hoja['A1'] = 'DESGLOSE HORAS HOMBRE'
        hoja['A1'].font = self.estilo_titulo
        hoja.merge_cells('A1:F1')
        
        # Encabezados
        encabezados = ['Máquina', 'Fecha', 'Mecánico', 'Tipo Orden', 'Horas', 'Costo ($35.000/h)']
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=3, column=col)
            celda.value = encabezado
            self._aplicar_estilo_encabezado(celda)
        
        fila = 4
        for hh in sorted(horas_hombre, key=lambda h: (h.codigo_maquina, h.fecha)):
            hoja.cell(row=fila, column=1, value=hh.codigo_maquina)
            hoja.cell(row=fila, column=2, value=hh.fecha.strftime('%d/%m/%Y'))
            hoja.cell(row=fila, column=3, value=hh.mecanico)
            hoja.cell(row=fila, column=4, value=hh.tipo_orden)
            hoja.cell(row=fila, column=5, value=float(hh.horas))
            hoja.cell(row=fila, column=6, value=self._formatear_moneda(hh.horas * Decimal('35000')))
            
            for col in range(1, 7):
                self._aplicar_borde(hoja.cell(row=fila, column=col))
            
            fila += 1
        
        # Ajustar ancho de columnas
        hoja.column_dimensions['A'].width = 20
        hoja.column_dimensions['B'].width = 12
        hoja.column_dimensions['C'].width = 30
        hoja.column_dimensions['D'].width = 15
        for col in range(5, 7):
            hoja.column_dimensions[get_column_letter(col)].width = 18
