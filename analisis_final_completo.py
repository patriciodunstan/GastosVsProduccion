import openpyxl
from decimal import Decimal

wb = openpyxl.load_workbook(r'C:\Users\patricio dunstan sae\GastosVsProduccion\informe_produccion_gastos.xlsx')

def parse_valor(cell):
    if cell is None:
        return Decimal('0')
    if isinstance(cell.value, (int, float)):
        return Decimal(str(cell.value))
    if isinstance(cell.value, str):
        limpio = cell.value.replace('$', '').replace('.', '').replace(',', '')
        return Decimal(limpio) if limpio else Decimal('0')
    return Decimal('0')

print('=== ANÁLISIS COMPLETO DEL INFORME FINAL ===\n')

# 1. ANÁLISIS DE RESUMEN TRIMESTRAL COMPLETO
print('1. ANÁLISIS DE PRODUCCIÓN VS GASTOS TRIMESTRAL')
print('=' * 100)

ws = wb['Resumen Trimestral Completo']

# Leer datos de la hoja de resumen
datos_maquinas = []

# Encabezados en fila 4
for row in ws.iter_rows(min_row=5, values_only=True):
    if row[0]:  # Si hay nombre de máquina
        maquina = row[0]
        prod_mt3 = parse_valor(type('Cell', (), {'value': row[1]})())
        prod_h = parse_valor(type('Cell', (), {'value': row[2]})())
        prod_km = parse_valor(type('Cell', (), {'value': row[3]})())
        prod_vueltas = parse_valor(type('Cell', (), {'value': row[4]})())
        prod_neta = parse_valor(type('Cell', (), {'value': row[5]})())
        
        # Gastos por categoría (columnas 7-22)
        gastos = {}
        gastos['repuestos'] = parse_valor(type('Cell', (), {'value': row[7]})())
        gastos['combustibles'] = parse_valor(type('Cell', (), {'value': row[8]})())
        gastos['reparaciones'] = parse_valor(type('Cell', (), {'value': row[9]})())
        gastos['seguros'] = parse_valor(type('Cell', (), {'value': row[10]})())
        gastos['honorarios'] = parse_valor(type('Cell', (), {'value': row[11]})())
        gastos['epp'] = parse_valor(type('Cell', (), {'value': row[12]})())
        gastos['peajes'] = parse_valor(type('Cell', (), {'value': row[13]})())
        gastos['remuneraciones'] = parse_valor(type('Cell', (), {'value': row[14]})())
        gastos['permisos'] = parse_valor(type('Cell', (), {'value': row[15]})())
        gastos['alimentacion'] = parse_valor(type('Cell', (), {'value': row[16]})())
        gastos['pasajes'] = parse_valor(type('Cell', (), {'value': row[17]})())
        gastos['correspondencia'] = parse_valor(type('Cell', (), {'value': row[18]})())
        gastos['gastos_legales'] = parse_valor(type('Cell', (), {'value': row[19]})())
        gastos['multas'] = parse_valor(type('Cell', (), {'value': row[20]})())
        gastos['otros_gastos'] = parse_valor(type('Cell', (), {'value': row[21]})())
        gastos['total_op'] = parse_valor(type('Cell', (), {'value': row[22]})())
        
        total_gastos = sum(gastos.values())
        
        # Resultado (columna 26 según el código)
        try:
            resultado = parse_valor(type('Cell', (), {'value': row[25]})())
        except:
            resultado = prod_neta - total_gastos  # Calcular manualmente si falla
        
        datos_maquinas.append({
            'maquina': maquina,
            'prod_mt3': prod_mt3,
            'prod_h': prod_h,
            'prod_km': prod_km,
            'prod_vueltas': prod_vueltas,
            'prod_neta': prod_neta,
            'gastos': gastos,
            'total_gastos': total_gastos,
            'resultado': resultado
        })

# 2. MÁQUINAS CON PEOR RELACIÓN PRODUCCIÓN VS GASTO
print('\n2. TOP 20 MÁQUINAS CON MAYOR GASTO RELATIVO (Producción Neta vs Gasto Total)')
print('=' * 100)
print('{:<40} {:>12} {:>15} {:>15} {:>15} {:>15}'.format(
    'MÁQUINA', 'Prod Neta', 'Total Gastos', 'Ratio', 'Resultado', 'Margen %'
))
print('-' * 100)

relaciones = []
for d in datos_maquinas:
    if d['prod_neta'] > 0 and d['total_gastos'] > 0:
        ratio = d['total_gastos'] / d['prod_neta']
        resultado = d['resultado']
        margen = (resultado / d['prod_neta']) * 100
        relaciones.append((d['maquina'], d['prod_neta'], d['total_gastos'], ratio, resultado, margen))

# Ordenar por ratio de mayor a peor
relaciones.sort(key=lambda x: x[3], reverse=True)

for i, (maquina, prod_neta, total_gastos, ratio, resultado, margen) in enumerate(relaciones[:20], 1):
    print('{:<40} {:>12} {:>15} {:>15.2f} {:>15} {:>15.1f}%'.format(
        maquina[:40],
        f'${prod_neta:,.0f}',
        f'${total_gastos:,.0f}',
        ratio,
        f'${resultado:,.0f}',
        margen
    ))

# 3. MÁQUINAS CON MAYOR GASTO TOTAL
print('\n3. TOP 20 MÁQUINAS CON MAYOR GASTO TOTAL')
print('=' * 100)
print('{:<40} {:>15} {:>15} {:>15} {:>15}'.format(
    'MÁQUINA', 'Total Gastos', 'Repuestos', 'Combustibles', 'Reparaciones', 'Resultado'
))
print('-' * 100)

por_gasto = sorted(datos_maquinas, key=lambda x: x['total_gastos'], reverse=True)[:20]
for d in por_gasto:
    print('{:<40} {:>15} {:>15} {:>15} {:>15} {:>15}'.format(
        d['maquina'][:40],
        f'${d["total_gastos"]:,.0f}',
        f'${d["gastos"]["repuestos"]:,.0f}',
        f'${d["gastos"]["combustibles"]:,.0f}',
        f'${d["gastos"]["reparaciones"]:,.0f}',
        f'${d["resultado"]:,.0f}'
    ))

# 4. MÁQUINAS CON PEOR RESULTADO (más pérdidas)
print('\n4. TOP 20 MÁQUINAS CON PEOR RESULTADO (MÁS PÉRDIDAS)')
print('=' * 100)
print('{:<40} {:>12} {:>15} {:>15} {:>15} {:>15}'.format(
    'MÁQUINA', 'Prod Neta', 'Gastos', 'Resultado', 'Prod H', 'Prod MT3'
))
print('-' * 100)

por_resultado = sorted([d for d in datos_maquinas if d['resultado'] < 0], key=lambda x: x['resultado'])[:20]
for d in por_resultado:
    print('{:<40} {:>12} {:>15} {:>15} {:>15} {:>15}'.format(
        d['maquina'][:40],
        f'${d["prod_neta"]:,.0f}',
        f'${d["total_gastos"]:,.0f}',
        f'${d["resultado"]:,.0f}',
        f'{d["prod_h"]:,.0f}',
        f'{d["prod_mt3"]:,.0f}'
    ))

# 5. ANÁLISIS DE CATEGORÍAS DE GASTO
print('\n5. TOP 10 CATEGORÍAS DE GASTO POR MONTO TOTAL')
print('=' * 100)

totales_categoria = {
    'Repuestos': Decimal('0'),
    'Combustibles': Decimal('0'),
    'Reparaciones': Decimal('0'),
    'Seguros': Decimal('0'),
    'Honorarios': Decimal('0'),
    'EPP': Decimal('0'),
    'Peajes': Decimal('0'),
    'Remuneraciones': Decimal('0'),
    'Permisos': Decimal('0'),
    'Alimentación': Decimal('0'),
    'Pasajes': Decimal('0'),
    'Correspondencia': Decimal('0'),
    'Gastos Legales': Decimal('0'),
    'Multas': Decimal('0'),
    'Otros': Decimal('0')
}

for d in datos_maquinas:
    totales_categoria['Repuestos'] += d['gastos']['repuestos']
    totales_categoria['Combustibles'] += d['gastos']['combustibles']
    totales_categoria['Reparaciones'] += d['gastos']['reparaciones']
    totales_categoria['Seguros'] += d['gastos']['seguros']
    totales_categoria['Honorarios'] += d['gastos']['honorarios']
    totales_categoria['EPP'] += d['gastos']['epp']
    totales_categoria['Peajes'] += d['gastos']['peajes']
    totales_categoria['Remuneraciones'] += d['gastos']['remuneraciones']
    totales_categoria['Permisos'] += d['gastos']['permisos']
    totales_categoria['Alimentación'] += d['gastos']['alimentacion']
    totales_categoria['Pasajes'] += d['gastos']['pasajes']
    totales_categoria['Correspondencia'] += d['gastos']['correspondencia']
    totales_categoria['Gastos Legales'] += d['gastos']['gastos_legales']
    totales_categoria['Multas'] += d['gastos']['multas']
    totales_categoria['Otros'] += d['gastos']['otros_gastos']

total_general = sum(totales_categoria.values())

categorias_ordenadas = sorted(totales_categoria.items(), key=lambda x: x[1], reverse=True)
for categoria, monto in categorias_ordenadas:
    porcentaje = (monto / total_general) * 100
    print('{:<25} {:>20} {:>10.1f}%'.format(
        categoria,
        f'${monto:,.0f}',
        porcentaje
    ))

print('=' * 100)
print('{:<25} {:>20}'.format('TOTAL GENERAL', f'${total_general:,.0f}'))

# 6. MÁQUINAS EFICIENTES (mejor resultado)
print('\n6. TOP 20 MÁQUINAS MÁS EFICIENTES (MEJOR RESULTADO)')
print('=' * 100)
print('{:<40} {:>12} {:>15} {:>15} {:>15}'.format(
    'MÁQUINA', 'Prod Neta', 'Gastos', 'Resultado', 'Margen %'
))
print('-' * 100)

eficientes = sorted([d for d in datos_maquinas if d['resultado'] > 0], key=lambda x: x['resultado'], reverse=True)[:20]
for d in eficientes:
    margen = (d['resultado'] / d['prod_neta']) * 100 if d['prod_neta'] > 0 else 0
    print('{:<40} {:>12} {:>15} {:>15} {:>15.1f}%'.format(
        d['maquina'][:40],
        f'${d["prod_neta"]:,.0f}',
        f'${d["total_gastos"]:,.0f}',
        f'${d["resultado"]:,.0f}',
        margen
    ))

# 7. RESUMEN EJECUTIVO
print('\n7. RESUMEN EJECUTIVO')
print('=' * 100)

total_prod_neta = sum(d['prod_neta'] for d in datos_maquinas)
total_gastos = sum(d['total_gastos'] for d in datos_maquinas)
total_resultado = sum(d['resultado'] for d in datos_maquinas)

maquinas_perdidas = sum(1 for d in datos_maquinas if d['resultado'] < 0)
maquinas_ganancia = sum(1 for d in datos_maquinas if d['resultado'] > 0)
maquinas_neutras = sum(1 for d in datos_maquinas if d['resultado'] == 0)

print(f'Total Producción Neta:    ${total_prod_neta:,.0f}')
print(f'Total Gastos:            ${total_gastos:,.0f}')
print(f'Total Resultado:          ${total_resultado:,.0f}')
print(f'Margen Promedio:         {(total_resultado / total_prod_neta * 100) if total_prod_neta > 0 else 0:.1f}%')
print()
print(f'Máquinas con pérdida:    {maquinas_perdidas}')
print(f'Máquinas con ganancia:   {maquinas_ganancia}')
print(f'Máquinas neutras:        {maquinas_neutras}')
print(f'Total máquinas:          {len(datos_maquinas)}')

# 8. ANÁLISIS DE GASTOS POR MES
print('\n8. ANÁLISIS DE GASTOS POR MES')
print('=' * 100)

ws_detalle = wb['Detalle Gastos Completo']

gastos_por_mes = {10: Decimal('0'), 11: Decimal('0'), 12: Decimal('0')}
maquina_mes_map = {10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}

for row in ws_detalle.iter_rows(min_row=4, values_only=True):
    if row[0] and row[1]:
        mes = row[1]  # Mes está en columna 1
        if isinstance(mes, int):
            total = parse_valor(type('Cell', (), {'value': row[17]})())  # Total en columna 17
            gastos_por_mes[mes] += total

for mes, monto in gastos_por_mes.items():
    print(f'{maquina_mes_map[mes]}: ${monto:,.0f}')

print('=' * 100)
print(f'TOTAL: ${sum(gastos_por_mes.values()):,.0f}')

# 9. IDENTIFICAR MÁQUINAS PROBLEMÁTICAS (alto gasto, baja producción)
print('\n9. MÁQUINAS CRÍTICAS - Alto Gasto + Baja Producción')
print('=' * 100)
print('(Máquinas con gastos > $100M y producción neta < $50M)')
print()

criticas = [d for d in datos_maquinas if d['total_gastos'] > Decimal('100000000') and d['prod_neta'] < Decimal('50000000')]
criticas.sort(key=lambda x: x['total_gastos'], reverse=True)

if criticas:
    print('{:<40} {:>15} {:>15} {:>15} {:>15}'.format(
        'MÁQUINA', 'Gastos', 'Prod Neta', 'Resultado', 'Ratio G/Prod'
    ))
    print('-' * 100)
    
    for d in criticas:
        ratio = d['total_gastos'] / d['prod_neta'] if d['prod_neta'] > 0 else 0
        print('{:<40} {:>15} {:>15} {:>15} {:>15.2f}x'.format(
            d['maquina'][:40],
            f'${d["total_gastos"]:,.0f}',
            f'${d["prod_neta"]:,.0f}',
            f'${d["resultado"]:,.0f}',
            ratio
        ))
else:
    print('No se encontraron máquinas que cumplan el criterio de alto gasto + baja producción.')

print('\n=== FIN DEL ANÁLISIS ===')
