import openpyxl
from decimal import Decimal

wb = openpyxl.load_workbook(r'C:\Users\patricio dunstan sae\GastosVsProduccion\informe_produccion_gastos.xlsx')

def parse_valor(cell):
    if cell is None:
        return Decimal('0')
    if isinstance(cell.value, (int, float)):
        return Decimal(str(cell.value))
    if isinstance(cell.value, str):
        # Remover símbolo $ y puntos de miles
        limpio = cell.value.replace('$', '').replace('.', '').replace(',', '')
        if limpio.isdigit():
            return Decimal(limpio)
    return Decimal('0')

print('=== ANÁLISIS COMPLETO DE LOS DATOS DEL EXCEL ===')

# 1. REPUESTOS (DESGLOSE REPUESTOS)
print('\n1. REPUESTOS DE DATABODEGA (Q4 2025):')
ws_repuestos = wb['Desglose Repuestos']

total_repuestos = Decimal('0')
maquinas_con_repuestos = set()

for row in ws_repuestos.iter_rows(min_row=4, values_only=True):
    if row[0]:  # Si hay nombre de máquina
        total = parse_valor(type('Cell', (), {'value': row[5]})())
        if total > 0:
            maquinas_con_repuestos.add(row[0])
            total_repuestos += total

print(f'   Total repuestos: ${total_repuestos:,.0f}')
print(f'   Máquinas con repuestos: {len(maquinas_con_repuestos)}')

# 2. GASTOS OPERACIONALES (DESAGLOSE GASTOS POR TIPO)
print('\n2. GASTOS OPERACIONALES (REPORTES CONTABLES):')
print('   Filtrando solo cuentas de gastos (401xxx)...')

ws_gastos = wb['Desglose Gastos por Tipo']

# Columnas: 0=Cod Maquina, 1=Mes, 2=Tipo Gasto, 6=Origen, 7=Monto
total_gastos_op = Decimal('0')
tipos_gastos = {}

for row in ws_gastos.iter_rows(min_row=4, values_only=True):
    if row[0]:  # Si hay código de máquina
        tipo_gasto = str(row[2]) if row[2] else ''
        origen = str(row[6]) if row[6] else ''
        
        # Solo considerar gastos (cuentas 401xxx)
        if tipo_gasto.startswith('401'):
            monto = parse_valor(type('Cell', (), {'value': row[7]})())
            
            if monto > 0:
                total_gastos_op += monto
                
                if tipo_gasto not in tipos_gastos:
                    tipos_gastos[tipo_gasto] = {'count': 0, 'total': Decimal('0')}
                tipos_gastos[tipo_gasto]['count'] += 1
                tipos_gastos[tipo_gasto]['total'] += monto

print(f'   TOTAL GASTOS OPERACIONALES: ${total_gastos_op:,.0f}')
print(f'   Tipos de gastos encontrados: {len(tipos_gastos)}')

# Mostrar top 5 tipos de gastos
print('\n   TOP 5 TIPOS DE GASTOS POR MONTO:')
for tipo, data in sorted(tipos_gastos.items(), key=lambda x: x[1]['total'], reverse=True)[:5]:
    print(f'      {tipo:15}: ${data["total"]:,.0f} ({data["count"]} registros)')

# 3. DETALLE GASTOS POR MES Y CATEGORÍA
print('\n3. DESGLOSE POR CATEGORÍA (DESGLOSE GASTOS POR MES):')
ws_detalle = wb['Detalle Gastos Completo']

# Columnas: 0=Máquina, 2=Repuestos, 3=Combustibles, 4=Reparaciones, 5=Seguros, 6=Honorarios,
#           7=EPP, 8=Peajes, 9=Remuneraciones, 10=Permisos, 11=Alimentación, 12=Pasajes,
#           13=Correspondencia, 14=Gastos Legales, 15=Multas, 16=Otros, 17=Total Gastos

categorias_detalle = {
    'Repuestos': 2,
    'Combustibles': 3,
    'Reparaciones': 4,
    'Seguros': 5,
    'Honorarios': 6,
    'EPP': 7,
    'Peajes': 8,
    'Remuneraciones': 9,
    'Permisos': 10,
    'Alimentacion': 11,
    'Pasajes': 12,
    'Correspondencia': 13,
    'Gastos Legales': 14,
    'Multas': 15,
    'Otros': 16,
    'Total Gastos': 17
}

for nombre, col_idx in categorias_detalle.items():
    categoria_total = Decimal('0')
    for row in ws_detalle.iter_rows(min_row=4, values_only=True):
        if row[0]:
            monto = parse_valor(type('Cell', (), {'value': row[col_idx]})())
            categoria_total += monto
    print(f'   {nombre:20}: ${categoria_total:,.0f}')

# 4. RESUMEN GENERAL
print('\n=== RESUMEN GENERAL DE GASTOS (Q4 2025) ===')
total_general = total_repuestos + total_gastos_op
print(f'   Repuestos (DATABODEGA):       ${total_repuestos:,.0f}')
print(f'   Gastos Operacionales (401xxx): ${total_gastos_op:,.0f}')
print('-' * 60)
print(f'   TOTAL GENERAL GASTOS:          ${total_general:,.0f}')

print('\n=== COMPARACIÓN CON VALORES ESPERADOS ===')
print('   Valores reportados en análisis:')
print('   - Repuestos Q4 2025:     $171.189.015')
print('   - Gastos operacionales:    $700.962.517 (solo 401xxx)')
print(f'\n   Valores en el Excel:')
print(f'   - Repuestos:              ${total_repuestos:,.0f}')
print(f'   - Gastos operacionales:    ${total_gastos_op:,.0f}')
print(f'\n   Diferencia Repuestos:      ${total_repuestos - Decimal("171189015"):,.0f}')
print(f'   Diferencia Gastos Op:     ${total_gastos_op - Decimal("700962517"):,.0f}')
